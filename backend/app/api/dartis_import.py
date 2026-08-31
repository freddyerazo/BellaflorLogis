"""
API para importar archivos Excel de Dartis desde el frontend.

POST /api/dartis/upload
  - file_recetas : Excel de Ventas Recetas (obligatorio)
  - file_ventas  : Excel de Ventas clasico (obligatorio)

Regla: se suben los dos archivos juntos. Primero se procesa Recetas
(inserta registros), luego Ventas (enriquece vendedor y agencia_carga).
"""

import asyncio
import logging
import tempfile
from datetime import date
from pathlib import Path

import openpyxl
from fastapi import APIRouter, File, HTTPException, UploadFile
from psycopg2.extras import execute_values
from sqlalchemy import text

from app.database.connection import engine

router = APIRouter(prefix="/dartis", tags=["Dartis Import"])
logger = logging.getLogger(__name__)

RECETAS_DATA_START = 7
VENTAS_DATA_START  = 7


# -- Helpers ------------------------------------------------------------------
def _safe_str(val):
    return str(val).strip() if val is not None else None

def _safe_int(val):
    try: return int(val)
    except: return None

def _safe_float(val):
    try: return float(val)
    except: return None

def _safe_date(val):
    """Normaliza a date. Dartis no siempre exporta la columna como fecha de
    Excel; cuando viene como texto, devolverla tal cual rompia el import:
    la reconciliacion terminaba comparando date = text, que en Postgres no
    existe como operador, y abortaba toda la transaccion."""
    if val is None: return None
    if hasattr(val, "date"): return val.date()
    if isinstance(val, str):
        try:
            # admite "2026-08-24" y "2026-08-24 00:00:00"
            return date.fromisoformat(val.strip()[:10])
        except ValueError:
            return val
    return val

def _norm_encabezado(v):
    """Normaliza un encabezado para compararlo: sin espacios, tildes ni mayusculas."""
    import unicodedata
    t = str(v or "").strip().lower().replace(" ", "").replace("_", "")
    t = unicodedata.normalize("NFD", t)
    return "".join(c for c in t if unicodedata.category(c) != "Mn")


def _mapear_columnas(rows, alias: dict, clave_obligatoria: str):
    """Ubica las columnas por NOMBRE de encabezado, no por posicion.

    Dartis agrega columnas al archivo sin avisar: cuando aparecio `paisVenta`
    en la sexta columna, el vendedor se corrio un lugar y el importador, que
    leia por indice fijo, habria escrito el nombre del pais dentro de
    `vendedor` en las 24.000 filas. Leer por encabezado hace que agregar o
    reordenar columnas deje de romper la importacion.

    Devuelve (indices, fila_datos) o (None, 0) si no encuentra el encabezado.
    """
    for i, row in enumerate(rows[:20]):
        # El encabezado de Dartis viene partido en dos filas: la primera trae
        # los nombres de campo y un "total" generico, y la siguiente los
        # subtitulos de ese total (piezas / tallos / dolares). Se buscan los
        # alias en ambas.
        norm = [_norm_encabezado(v) for v in row]

        # Primero solo con esta fila. La clave obligatoria TIENE que estar
        # aca: si se aceptara encontrarla en la fila siguiente, una fila vacia
        # anterior al encabezado pasaria por encabezado y devolveria un mapeo
        # incompleto (justamente lo que pasaba con los totales).
        indices = {}
        for campo, nombres in alias.items():
            for n in nombres:
                if n in norm:
                    indices[campo] = norm.index(n)
                    break

        if clave_obligatoria in indices:
            # Recien ahora se completan los que faltan con la fila siguiente:
            # el encabezado de Dartis viene partido en dos, con un "total"
            # generico arriba y sus subtitulos (piezas/tallos/dolares) abajo.
            sig = [_norm_encabezado(v) for v in rows[i + 1]] if i + 1 < len(rows) else []
            for campo, nombres in alias.items():
                if campo in indices:
                    continue
                for n in nombres:
                    if n in sig:
                        indices[campo] = sig.index(n)
                        break

            # Los datos empiezan en la primera fila posterior cuya clave sea
            # numerica: entre el encabezado y los datos puede haber una fila
            # de subtitulos (en Ventas, "Piezas"/"dolares").
            for j in range(i + 1, len(rows)):
                if _safe_int(rows[j][indices[clave_obligatoria]]) is not None:
                    return indices, j
    return None, 0


def _load_wb(upload: UploadFile):
    """Guarda el archivo en un temporal y abre el workbook."""
    suffix = Path(upload.filename).suffix or ".xlsx"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(upload.file.read())
    tmp.flush()
    return openpyxl.load_workbook(tmp.name, data_only=True, read_only=True)

def _load_wb_bytes(data: bytes, filename: str):
    """Carga un workbook desde bytes ya leídos."""
    suffix = Path(filename).suffix or ".xlsx"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(data)
    tmp.flush()
    return openpyxl.load_workbook(tmp.name, data_only=True, read_only=True)


# -- Logica importacion -------------------------------------------------------
BATCH_SIZE = 500

# Encabezados del archivo Ventas Recetas. Igual que en Ventas, se ubican por
# nombre y no por posicion: cuando Dartis agrego `variedad_receta` la inserto
# en el medio (posicion 9), corriendo guia_madre, guia_hija, tipo_caja y los
# tres totales un lugar a la derecha. Leyendo por indice fijo, tipo_caja
# ("QB") habria terminado en total_piezas y los tallos en los dolares.
RECETAS_ALIAS = {
    "fecha":         ("fecha",),
    "dae":           ("dae",),
    "id_com":        ("idcomercializadora",),
    "id_pedido":     ("idpedido",),
    "empresa":       ("empresa",),
    "cliente":       ("cliente",),
    "destinatario":  ("destinatario",),
    "postcosecha":   ("postcosecha",),
    "especie":       ("especie",),
    "variedad":      ("variedadreceta", "variedad"),
    "guia_madre":    ("guiamadre",),
    "guia_hija":     ("guiahija",),
    "tipo_caja":     ("tipocaja",),
    "total_piezas":  ("piezas",),
    "total_tallos":  ("tallos",),
    "total_dolares": ("dolares",),
}


def _import_recetas(ws, conn) -> dict:
    todas = list(ws.iter_rows(values_only=True))
    idx, inicio = _mapear_columnas(todas, RECETAS_ALIAS, "id_pedido")

    if idx is None:
        idx = {"fecha": 0, "dae": 1, "id_com": 2, "id_pedido": 3, "empresa": 4,
               "cliente": 5, "destinatario": 6, "postcosecha": 7, "especie": 8,
               "guia_madre": 9, "guia_hija": 10, "tipo_caja": 11,
               "total_piezas": 12, "total_tallos": 13, "total_dolares": 14}
        inicio = RECETAS_DATA_START
        logger.warning("Recetas: no se reconocio el encabezado, se usan posiciones fijas")

    rows = todas[inicio:]
    postcosechas = set()
    params = []
    errores = 0

    def col(row, campo):
        i = idx.get(campo)
        return row[i] if i is not None and i < len(row) else None

    for row in rows:
        if not any(row):
            continue
        try:
            id_pedido = _safe_int(col(row, "id_pedido"))
            if not id_pedido:
                continue
            pc = _safe_str(col(row, "postcosecha"))
            if pc:
                postcosechas.add(pc)
            params.append({
                "fecha":        _safe_date(col(row, "fecha")),
                "dae":          _safe_str(col(row, "dae")),
                "id_com":       _safe_int(col(row, "id_com")),
                "id_pedido":    id_pedido,
                "empresa":      _safe_str(col(row, "empresa")),
                "cliente":      _safe_str(col(row, "cliente")),
                "destinatario": _safe_str(col(row, "destinatario")),
                "postcosecha":  pc,
                "especie":      _safe_str(col(row, "especie")),
                "variedad":     _safe_str(col(row, "variedad")),
                "guia_madre":   _safe_str(col(row, "guia_madre")),
                "guia_hija":    _safe_str(col(row, "guia_hija")),
                "tipo_caja":    _safe_str(col(row, "tipo_caja")),
                "total_piezas": _safe_float(col(row, "total_piezas")),
                "total_tallos": _safe_int(col(row, "total_tallos")),
                "total_dolares":_safe_float(col(row, "total_dolares")),
            })
        except Exception:
            errores += 1

    # Agrupar por clave única (incluye especie: una misma guia puede
    # transportar varias especies distintas). Si dos lineas comparten
    # la clave completa (mismo pedido+guia+caja+especie), son lotes
    # separados del mismo producto y se suman sus cantidades.
    # La clave unica no incluye la variedad, y una misma linea puede agrupar
    # varias: en el archivo del 2026-08-30, 702 de 2.017 claves traen mas de
    # una, y un pedido de BOUQUETS llega a 53 por ser producto compuesto. Se
    # conservan TODAS en una lista en vez de quedarse con una sola, que daria
    # un dato preciso en apariencia y equivocado en el fondo.
    dedup = {}
    variedades_por_clave = {}
    for p in params:
        key = (p["id_pedido"], p["guia_madre"], p["guia_hija"], p["tipo_caja"], p["especie"])
        if p.get("variedad"):
            variedades_por_clave.setdefault(key, set()).add(p["variedad"])
        if key in dedup:
            existing = dedup[key]
            existing["total_piezas"] = (existing["total_piezas"] or 0) + (p["total_piezas"] or 0)
            existing["total_tallos"] = (existing["total_tallos"] or 0) + (p["total_tallos"] or 0)
            existing["total_dolares"] = (existing["total_dolares"] or 0) + (p["total_dolares"] or 0)
        else:
            dedup[key] = p

    for key, p in dedup.items():
        vs = variedades_por_clave.get(key)
        p["variedad_receta"] = ", ".join(sorted(vs)) if vs else None

    params = list(dedup.values())

    tuples = [
        (p["fecha"], p["dae"], p["id_com"], p["id_pedido"],
         p["empresa"], p["cliente"], p["destinatario"], p["postcosecha"], p["especie"],
         p["variedad_receta"],
         p["guia_madre"], p["guia_hija"], p["tipo_caja"],
         p["total_piezas"], p["total_tallos"], p["total_dolares"])
        for p in params
    ]

    raw = conn.connection.cursor()

    # Marca de tiempo ANTES de tocar nada: cualquier fila que no se toque en
    # este import (ni se inserte ni se actualice) se queda con un
    # importado_at anterior a esta marca -- eso es lo que distingue "ya no
    # esta en el archivo nuevo" de "sigue igual".
    inicio = conn.execute(text("SELECT now()")).scalar()

    # Dartis a veces exporta una linea sin guia (todavia no asignada) y luego,
    # en una reimportacion posterior, la misma linea ya con guia. Como la guia
    # es parte de la clave unica, ON CONFLICT no reconoce esa fila como la
    # misma -- se duplicaria. Se completa la guia en la fila vieja ANTES del
    # upsert principal, para que el ON CONFLICT de mas abajo si la reconozca.
    con_guia = [p for p in params if p["guia_madre"] or p["guia_hija"]]
    if con_guia:
        raw.execute("""
            CREATE TEMP TABLE tmp_recetas_guias (
                id_pedido INTEGER, tipo_caja TEXT, especie TEXT, fecha DATE,
                guia_madre TEXT, guia_hija TEXT
            ) ON COMMIT DROP
        """)
        execute_values(raw, "INSERT INTO tmp_recetas_guias VALUES %s", [
            (p["id_pedido"], p["tipo_caja"], p["especie"], p["fecha"], p["guia_madre"], p["guia_hija"])
            for p in con_guia
        ], page_size=2000)
        # dv.id = (subquery LIMIT 1), no un UPDATE...FROM directo: si existen
        # DOS lineas viejas sin guia para la misma clave (dos lotes separados
        # del mismo producto que Dartis aun no distinguia), un UPDATE...FROM
        # les pondria a AMBAS la misma guia nueva -- duplicando la clave unica.
        # Con la subquery solo se completa una fila por cada linea del archivo
        # nuevo; la otra queda para una importacion futura que si la distinga.
        #
        # El NOT EXISTS es necesario porque la tabla tiene ADEMAS un indice
        # unico "dartis_ventas_clave_linea" con NULLS NOT DISTINCT (no creado
        # por este codigo) -- si la guia nueva ya pertenece a otra fila
        # existente, completar la fila sin guia con esa misma guia duplicaria
        # esa clave y tumbaba la importacion entera (visto en produccion:
        # pedido 201520/LYSIMACHIA). En ese caso se deja la fila sin guia
        # -- el upsert principal de mas abajo la vuelve a intentar con sus
        # valores reales, y como esos SI son valores no nulos, ON CONFLICT
        # la reconoce y actualiza en vez de duplicar.
        raw.execute("""
            UPDATE dartis_ventas dv
            SET    guia_madre = t.guia_madre, guia_hija = t.guia_hija
            FROM   tmp_recetas_guias t
            WHERE  dv.id = (
                SELECT dv2.id FROM dartis_ventas dv2
                WHERE dv2.id_pedido = t.id_pedido AND dv2.tipo_caja = t.tipo_caja
                  AND dv2.especie = t.especie AND dv2.fecha = t.fecha
                  AND dv2.guia_madre IS NULL AND dv2.guia_hija IS NULL
                ORDER BY dv2.id
                LIMIT 1
            )
            AND NOT EXISTS (
                SELECT 1 FROM dartis_ventas dv3
                WHERE dv3.id_pedido = t.id_pedido AND dv3.tipo_caja = t.tipo_caja
                  AND dv3.especie = t.especie
                  AND dv3.guia_madre = t.guia_madre AND dv3.guia_hija = t.guia_hija
            )
        """)

    execute_values(raw, """
        INSERT INTO dartis_ventas (
            fecha, dae, id_comercializadora, id_pedido,
            empresa, cliente, destinatario, postcosecha, especie,
            variedad_receta,
            guia_madre, guia_hija, tipo_caja,
            total_piezas, total_tallos, total_dolares, active
        ) VALUES %s
        ON CONFLICT (id_pedido, guia_madre, guia_hija, tipo_caja, especie) DO UPDATE SET
            fecha               = EXCLUDED.fecha,
            dae                 = EXCLUDED.dae,
            id_comercializadora = EXCLUDED.id_comercializadora,
            empresa             = EXCLUDED.empresa,
            cliente             = EXCLUDED.cliente,
            destinatario        = EXCLUDED.destinatario,
            postcosecha         = EXCLUDED.postcosecha,
            especie             = EXCLUDED.especie,
            variedad_receta     = EXCLUDED.variedad_receta,
            total_piezas        = EXCLUDED.total_piezas,
            total_tallos        = EXCLUDED.total_tallos,
            total_dolares       = EXCLUDED.total_dolares,
            active              = true,
            importado_at        = now()
    """, [t + (True,) for t in tuples], page_size=1000)

    # Reconciliacion: cualquier fila de dartis_ventas cuya fecha caiga dentro
    # de las fechas de este archivo, pero que este import no toco (sigue con
    # importado_at anterior al inicio de esta corrida), ya no esta en la
    # fuente -- se inactiva en vez de dejarla viva para siempre (caso real:
    # pedido 200546, huerfano hasta que se borro a mano).
    fechas_archivo = sorted({p["fecha"] for p in params if p["fecha"]})
    inactivados = 0
    if fechas_archivo:
        r = conn.execute(text("""
            UPDATE dartis_ventas
            SET active = false
            WHERE fecha = ANY(CAST(:fechas AS date[])) AND importado_at < :inicio AND active = true
        """), {"fechas": fechas_archivo, "inicio": inicio})
        inactivados = r.rowcount

    sin_finca = _sync_postcosechas(conn, postcosechas)
    clientes_result = _sync_customers(conn, {p["cliente"] for p in params if p.get("cliente")})

    return {
        "inactivados": inactivados,
        "insertados_o_actualizados": len(tuples),
        "errores": errores,
        "postcosechas_sin_finca": sin_finca,
        "clientes_vinculados": clientes_result["vinculados"],
        "clientes_nuevos": clientes_result["nuevos"],
    }


# Nombres de encabezado del archivo Ventas. Se listan variantes por si Dartis
# los renombra: lo unico que no puede faltar es IdFactura.
VENTAS_ALIAS = {
    "id_pedido":     ("idfactura",),
    "agencia_carga": ("agenciacarga",),
    "vendedor":      ("vendedorpacking", "vendedor"),
    "pais_venta":    ("paisventa", "pais"),
}


# Dartis escribe algunos paises distinto que el catalogo de countries (que usa
# los nombres de Agrocalidad). Se declaran explicitos: adivinarlos con una
# busqueda difusa es justo lo que no conviene en un dato que despues se cruza
# contra requisitos fitosanitarios. Si aparece uno nuevo sin equivalencia, la
# importacion lo devuelve en `paises_sin_equivalencia` en vez de callarselo.
DARTIS_PAIS_ALIAS = {
    "bahrein": "Barein",
    "singapur": "Singapore",
    "paisesbajosholanda": "Paises Bajos",
}


def _enrich_ventas(ws, conn) -> dict:
    todas = list(ws.iter_rows(values_only=True))
    idx, inicio = _mapear_columnas(todas, VENTAS_ALIAS, "id_pedido")

    if idx is None:
        # Sin encabezado reconocible se vuelve a las posiciones historicas,
        # que son las del archivo anterior a que apareciera paisVenta.
        idx, inicio = {"id_pedido": 1, "agencia_carga": 4, "vendedor": 5}, VENTAS_DATA_START
        logger.warning("Ventas: no se reconocio el encabezado, se usan posiciones fijas")

    rows = todas[inicio:]
    agencias = set()
    params = []
    errores = 0

    def col(row, campo):
        i = idx.get(campo)
        return row[i] if i is not None and i < len(row) else None

    for row in rows:
        if not any(row):
            continue
        try:
            id_pedido = _safe_int(col(row, "id_pedido"))
            if not id_pedido:
                continue
            ag = _safe_str(col(row, "agencia_carga"))
            if ag:
                agencias.add(ag)
            params.append({
                "id_pedido":    id_pedido,
                "agencia_carga": ag,
                "vendedor":     _safe_str(col(row, "vendedor")),
                "pais_venta":   _safe_str(col(row, "pais_venta")),
            })
        except Exception:
            errores += 1

    # Un mismo id_pedido puede repetirse (embarques divididos). Sin
    # deduplicar, el UPDATE...FROM deja el resultado indefinido cuando
    # hay varias filas fuente para el mismo pedido. Se conserva la
    # última aparición en el archivo (mismo criterio que en recetas).
    dedup = {}
    for p in params:
        dedup[p["id_pedido"]] = p
    params = list(dedup.values())

    # El pais se resuelve aca y no en SQL: comparar sin tildes en Postgres
    # exigiria la extension unaccent, que no esta instalada en este proyecto.
    # Se prueba contra name_es (el nombre que usa Agrocalidad) y contra name.
    catalogo_paises = {}
    for cid, nombre_es, nombre in conn.execute(text(
            "SELECT id, name_es, name FROM countries")).all():
        for n in (nombre_es, nombre):
            if n:
                catalogo_paises.setdefault(_norm_encabezado(n), str(cid))

    sin_pais = set()
    for p in params:
        clave = _norm_encabezado(p["pais_venta"])
        if clave in DARTIS_PAIS_ALIAS:
            clave = _norm_encabezado(DARTIS_PAIS_ALIAS[clave])
        p["country_id"] = catalogo_paises.get(clave)
        if p["pais_venta"] and not p["country_id"]:
            sin_pais.add(p["pais_venta"])

    tuples = [(p["id_pedido"], p["agencia_carga"], p["vendedor"],
               p["pais_venta"], p["country_id"]) for p in params]

    raw = conn.connection.cursor()
    raw.execute("""
        CREATE TEMP TABLE tmp_ventas_enrich (
            id_pedido    INTEGER,
            agencia_carga TEXT,
            vendedor      TEXT,
            pais_venta    TEXT,
            country_id    UUID
        ) ON COMMIT DROP
    """)
    execute_values(raw, "INSERT INTO tmp_ventas_enrich VALUES %s", tuples, page_size=2000)
    # El texto crudo del pais se guarda aunque no matchee ningun pais del
    # catalogo, para poder re-resolver el mapeo sin volver a importar.
    raw.execute("""
        UPDATE dartis_ventas dv
        SET    vendedor      = t.vendedor,
               agencia_carga = t.agencia_carga,
               pais_venta    = t.pais_venta,
               country_id    = t.country_id
        FROM   tmp_ventas_enrich t
        WHERE  dv.id_pedido = t.id_pedido
    """)
    afectadas = raw.rowcount

    nuevas_ag = _sync_agencias(conn, agencias)

    return {
        "filas_procesadas": len(tuples),
        "actualizadas": afectadas,
        "errores": errores,
        "paises_sin_equivalencia": sorted(sin_pais),
        "agencias_nuevas": nuevas_ag,
    }


def _sync_agencias(conn, agencias: set) -> list:
    existentes = {
        r[0] for r in conn.execute(
            text("SELECT dartis_name FROM cargo_agencies WHERE dartis_name IS NOT NULL")
        )
    }
    agregadas = []
    for ag in sorted(agencias):
        if not ag or ag in existentes:
            continue
        base = "".join(c for c in ag if c.isalpha())[:3].upper()
        codigo = base
        n = 1
        while conn.execute(text("SELECT 1 FROM cargo_agencies WHERE code = :c"), {"c": codigo}).first():
            codigo = base[:2] + str(n)
            n += 1
        conn.execute(text("""
            INSERT INTO cargo_agencies (code, name, dartis_name, ocr_variants, type)
            VALUES (:code, :name, :dartis_name, '{}', 'aerea')
            ON CONFLICT (code) DO NOTHING
        """), {"code": codigo, "name": ag, "dartis_name": ag})
        agregadas.append(ag)
    return agregadas


def _sync_postcosechas(conn, postcosechas: set) -> list:
    existentes = {
        r[0].lower() for r in conn.execute(text("SELECT postcosecha FROM farm_postcosecha"))
    }
    return [pc for pc in sorted(postcosechas) if pc and pc.lower() not in existentes]


def _sync_customers(conn, clientes: set) -> dict:
    """Vincula clientes de Dartis con la tabla customers.
    - Busca por dartis_name (case-insensitive).
    - Crea automáticamente los que no existan (batch insert).
    - Actualiza customer_id en dartis_ventas con un solo UPDATE.
    """
    if not clientes:
        return {"vinculados": 0, "nuevos": 0}

    clientes = {cl for cl in clientes if cl}

    # Mapa dartis_name.lower() -> id de los ya existentes
    existentes = {
        r[0]: r[1]
        for r in conn.execute(text(
            "SELECT LOWER(dartis_name), id FROM customers WHERE dartis_name IS NOT NULL AND dartis_name != ''"
        ))
    }

    # Clientes que faltan
    faltantes = sorted(cl for cl in clientes if cl.lower() not in existentes)

    if faltantes:
        # Códigos existentes para evitar duplicados (carga en bulk)
        codigos_usados = {
            r[0] for r in conn.execute(text("SELECT customer_code FROM customers"))
        }

        nuevos_params = []
        for cl in faltantes:
            base = "".join(c for c in cl if c.isalnum())[:6].upper()
            codigo = base
            n = 1
            while codigo in codigos_usados:
                codigo = base[:5] + str(n)
                n += 1
            codigos_usados.add(codigo)
            nuevos_params.append({"code": codigo, "name": cl, "dartis_name": cl})

        raw = conn.connection.cursor()
        execute_values(raw, """
            INSERT INTO customers (customer_code, customer_name, dartis_name, active)
            VALUES %s
            ON CONFLICT DO NOTHING
        """, [(p["code"], p["name"], p["dartis_name"], True) for p in nuevos_params])

    # UPDATE masivo en dartis_ventas usando JOIN
    conn.execute(text("""
        UPDATE dartis_ventas dv
        SET    customer_id = c.id
        FROM   customers c
        WHERE  LOWER(TRIM(c.dartis_name)) = LOWER(TRIM(dv.cliente))
          AND  dv.customer_id IS DISTINCT FROM c.id
    """))

    vinculados = conn.execute(text(
        "SELECT COUNT(*) FROM dartis_ventas WHERE customer_id IS NOT NULL"
    )).scalar()

    return {"vinculados": vinculados, "nuevos": len(faltantes)}


# -- Endpoint -----------------------------------------------------------------
@router.post("/upload")
async def upload_dartis(
    file_recetas: UploadFile = File(..., description="Excel Ventas Recetas de Dartis"),
    file_ventas:  UploadFile = File(..., description="Excel Ventas clasico de Dartis"),
):
    """
    Sube los dos archivos Excel de Dartis en un solo paso:
    1. Procesa Ventas Recetas -> inserta/actualiza en dartis_ventas
    2. Procesa Ventas clasico -> enriquece vendedor y agencia_carga por id_pedido
    """
    contenido_recetas = await file_recetas.read()
    contenido_ventas  = await file_ventas.read()
    nombre_recetas    = file_recetas.filename
    nombre_ventas     = file_ventas.filename

    def _procesar():
        try:
            wb_recetas = _load_wb_bytes(contenido_recetas, nombre_recetas)
            wb_ventas  = _load_wb_bytes(contenido_ventas,  nombre_ventas)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error leyendo archivos: {e}")

        try:
            with engine.begin() as conn:
                resultado_recetas = _import_recetas(wb_recetas.active, conn)
                resultado_ventas  = _enrich_ventas(wb_ventas.active, conn)
        except Exception as e:
            # Antes esto se perdia como un 500 generico (Cloudflare/Render lo
            # muestra como pagina HTML, no JSON, asi que el frontend terminaba
            # mostrando el mensaje de respaldo "Error en el servidor" sin
            # ninguna pista de la causa real). Ahora queda logueado en el
            # servidor y el detalle real llega al frontend.
            logger.exception("Error importando Dartis (recetas=%s, ventas=%s)", nombre_recetas, nombre_ventas)
            raise HTTPException(status_code=500, detail=f"Error importando: {type(e).__name__}: {e}")

        return {
            "recetas": {"archivo": nombre_recetas, **resultado_recetas},
            "ventas":  {"archivo": nombre_ventas,  **resultado_ventas},
        }

    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _procesar)
