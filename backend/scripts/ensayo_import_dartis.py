# -*- coding: utf-8 -*-
"""Ensayo del importador de Dartis contra archivos reales, con ROLLBACK al final.

    cd backend
    C:\dev\venvs\blis\Scripts\python.exe scripts/ensayo_import_dartis.py

Ajustar las rutas BASE/RECETAS/VENTAS a los archivos que se quieran probar.

Ejecuta la logica real de _import_recetas y _enrich_ventas dentro de una
transaccion que se revierte, para verificar el formato nuevo sin escribir.
"""
import sys

import openpyxl
from sqlalchemy import text

import sys as _s
from pathlib import Path as _P
_s.path.insert(0, str(_P(__file__).resolve().parent.parent))

from app.api.dartis_import import _import_recetas, _enrich_ventas
from app.database.connection import engine

sys.stdout.reconfigure(encoding="utf-8")

BASE = r"C:/Users/coordinacion/OneDrive - BELLAFLOR GROUP INC/BFG/Downloads/"
RECETAS = BASE + "Ventas Recetas2026-09-03 21_47_33.xlsx"
VENTAS = BASE + "Ventas2026-09-03 21_46_31.xlsx"

wb_r = openpyxl.load_workbook(RECETAS, data_only=True, read_only=True)
wb_v = openpyxl.load_workbook(VENTAS, data_only=True, read_only=True)

conn = engine.connect()
trans = conn.begin()
try:
    antes = conn.execute(text("SELECT count(*) FROM dartis_ventas")).scalar()
    print(f"filas antes: {antes:,}\n")

    print("=== _import_recetas (formato nuevo, 16 columnas) ===")
    r = _import_recetas(wb_r[wb_r.sheetnames[0]], conn)
    for k, v in r.items():
        print(f"   {k}: {v}")

    print("\n=== _enrich_ventas (con paisVenta) ===")
    v = _enrich_ventas(wb_v[wb_v.sheetnames[0]], conn)
    for k, val in v.items():
        print(f"   {k}: {val}")

    despues = conn.execute(text("SELECT count(*) FROM dartis_ventas")).scalar()
    print(f"\nfilas despues: {despues:,}  (delta {despues - antes:+,})")

    print("\n=== como quedaron las variedades ===")
    for row in conn.execute(text("""
        SELECT especie, tipo_caja, variedad_receta,
               array_length(string_to_array(variedad_receta, ', '), 1) AS n
        FROM dartis_ventas
        WHERE variedad_receta IS NOT NULL
        ORDER BY n DESC NULLS LAST LIMIT 3
    """)).all():
        print(f"   {row[0]} / {row[1]} -> {row[3]} variedades")
        print(f"      {str(row[2])[:150]}...")

    print("\n   --- una linea de una sola variedad ---")
    for row in conn.execute(text("""
        SELECT especie, variedad_receta FROM dartis_ventas
        WHERE variedad_receta IS NOT NULL AND variedad_receta NOT LIKE '%, %'
        LIMIT 3
    """)).all():
        print(f"      {row[0]:16} -> {row[1]}")

    print("\n=== controles ===")
    checks = [
        ("filas con variedad", "SELECT count(*) FROM dartis_ventas WHERE variedad_receta IS NOT NULL"),
        ("tipo_caja que parece numero (corrupcion)",
         "SELECT count(*) FROM dartis_ventas WHERE tipo_caja ~ '^[0-9.]+$'"),
        ("guia_madre que parece variedad",
         "SELECT count(*) FROM dartis_ventas WHERE guia_madre ILIKE '%(b)%'"),
        ("total_tallos nulo con tallos esperados",
         "SELECT count(*) FROM dartis_ventas WHERE total_tallos IS NULL"),
        ("filas con pais resuelto",
         "SELECT count(*) FROM dartis_ventas WHERE country_id IS NOT NULL"),
    ]
    for etq, q in checks:
        print(f"   {etq:44} {conn.execute(text(q)).scalar():,}")

finally:
    trans.rollback()
    conn.close()
    print("\n>>> ROLLBACK: no se escribio nada en la base")
