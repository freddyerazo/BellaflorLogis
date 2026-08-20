"""
import_dartis.py — Importa y enriquece dartis_ventas desde los Excel de Dartis.

Comandos:
  import  — Importa Ventas Recetas a dartis_ventas (inserta nuevos, omite duplicados)
  enrich  — Rellena vendedor y agencia_carga desde el formato Ventas clásico (por id_pedido)

Uso:
  python scripts/import_dartis.py import  --file "data_dartis/VentasRecetas.xlsx"
  python scripts/import_dartis.py enrich  --file "data_dartis/Ventas.xlsx"

Requisito: backend/.env con DATABASE_URL configurado.
"""

import argparse
import os
import sys
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, text

# ── Config ────────────────────────────────────────────────────────────────────
BLIS_DIR = Path(__file__).resolve().parents[1]
load_dotenv(BLIS_DIR / "backend" / ".env")
DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    print("ERROR: DATABASE_URL no encontrado en backend/.env")
    sys.exit(1)

engine = create_engine(DATABASE_URL)

# Filas en el Excel de Dartis (0-based)
RECETAS_DATA_START  = 7   # Ventas Recetas: datos desde fila 8
VENTAS_DATA_START   = 7   # Ventas clasico: datos desde fila 8


# ── Helpers ───────────────────────────────────────────────────────────────────
def safe_str(val):
    return str(val).strip() if val else None

def safe_int(val):
    try: return int(val)
    except: return None

def safe_float(val):
    try: return float(val)
    except: return None

def safe_date(val):
    if val is None: return None
    if hasattr(val, 'date'): return val.date()
    return val


# ── Sync agencias nuevas ──────────────────────────────────────────────────────
def sync_agencias(conn, agencias: set) -> list:
    existentes = {
        row[0] for row in conn.execute(
            text("SELECT dartis_name FROM cargo_agencies WHERE dartis_name IS NOT NULL")
        )
    }
    agregadas = []
    for ag in sorted(agencias):
        if not ag or ag in existentes:
            continue
        base = ''.join(c for c in ag if c.isalpha())[:3].upper()
        codigo = base
        n = 1
        while conn.execute(text("SELECT 1 FROM cargo_agencies WHERE code = :c"), {"c": codigo}).first():
            codigo = base[:2] + str(n)
            n += 1
        conn.execute(text("""
            INSERT INTO cargo_agencies (code, name, dartis_name, ocr_variants, type)
            VALUES (:code, :name, :dartis_name, '{}', 'aerea')
            ON CONFLICT (code) DO NOTHING
        """), {"code": codigo, "name": ag, "dartis_name": ag})
        agregadas.append(ag)
    return agregadas


# ── Sync postcosechas nuevas ──────────────────────────────────────────────────
def sync_postcosechas(conn, postcosechas: set) -> list:
    existentes = {
        row[0].lower() for row in conn.execute(
            text("SELECT postcosecha FROM farm_postcosecha")
        )
    }
    return [pc for pc in sorted(postcosechas) if pc and pc.lower() not in existentes]


# ── COMANDO: import (Ventas Recetas) ─────────────────────────────────────────
def cmd_import(archivo: Path):
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[RECETAS_DATA_START:]

    agencias     = set()
    postcosechas = set()
    insertados   = 0
    duplicados   = 0
    errores      = 0

    with engine.begin() as conn:
        for row in data_rows:
            if not any(row):
                continue
            try:
                fecha        = safe_date(row[0])
                dae          = safe_str(row[1])
                id_com       = safe_int(row[2])
                id_pedido    = safe_int(row[3])
                empresa      = safe_str(row[4])
                cliente      = safe_str(row[5])
                destinatario = safe_str(row[6])
                postcosecha  = safe_str(row[7])
                especie      = safe_str(row[8])
                guia_madre   = safe_str(row[9])
                guia_hija    = safe_str(row[10])
                tipo_caja    = safe_str(row[11])
                total_piezas = safe_float(row[12])
                total_tallos = safe_int(row[13])
                total_dolares= safe_float(row[14])

                if postcosecha:
                    postcosechas.add(postcosecha)
                if not id_pedido:
                    continue

                result = conn.execute(text("""
                    INSERT INTO dartis_ventas (
                        fecha, dae, id_comercializadora, id_pedido,
                        empresa, cliente, destinatario, postcosecha, especie,
                        guia_madre, guia_hija, tipo_caja,
                        total_piezas, total_tallos, total_dolares
                    ) VALUES (
                        :fecha, :dae, :id_com, :id_pedido,
                        :empresa, :cliente, :destinatario, :postcosecha, :especie,
                        :guia_madre, :guia_hija, :tipo_caja,
                        :total_piezas, :total_tallos, :total_dolares
                    )
                    ON CONFLICT (id_pedido, guia_madre, guia_hija, tipo_caja) DO NOTHING
                """), {
                    "fecha": fecha, "dae": dae, "id_com": id_com, "id_pedido": id_pedido,
                    "empresa": empresa, "cliente": cliente, "destinatario": destinatario,
                    "postcosecha": postcosecha, "especie": especie,
                    "guia_madre": guia_madre, "guia_hija": guia_hija, "tipo_caja": tipo_caja,
                    "total_piezas": total_piezas, "total_tallos": total_tallos,
                    "total_dolares": total_dolares,
                })

                if result.rowcount > 0:
                    insertados += 1
                else:
                    duplicados += 1

            except Exception as e:
                errores += 1
                print(f"  AVISO fila omitida: {e}")

        print(f"  Registros insertados : {insertados}")
        print(f"  Duplicados omitidos  : {duplicados}")
        if errores:
            print(f"  Errores             : {errores}")

        nuevas_ag = sync_agencias(conn, agencias)
        print()
        if nuevas_ag:
            print("  Agencias nuevas agregadas:")
            for ag in nuevas_ag:
                print(f"    -> {ag}")
        else:
            print("  Agencias de carga: todas ya existian")

        sin_finca = sync_postcosechas(conn, postcosechas)
        if sin_finca:
            print("\n  AVISO postcosechas sin finca (asignar en farm_postcosecha):")
            for pc in sin_finca:
                print(f"    -> {pc}")
        else:
            print("  Postcosechas: todas mapeadas a una finca")


# ── COMANDO: enrich (Ventas clasico → vendedor + agencia_carga) ───────────────
def cmd_enrich(archivo: Path):
    """
    Lee el formato Ventas clasico de Dartis y actualiza dartis_ventas
    con vendedor y agencia_carga haciendo JOIN por id_pedido (= IdFactura).

    Columnas del Excel clasico (fila 5, 0-based):
      [0] Empresa  [1] IdFactura  [2] fechaSalida  [3] Postcosecha
      [4] agenciaCarga  [5] vendedorPacking  [6] total_piezas  [7] total_dolares
    """
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[VENTAS_DATA_START:]

    agencias   = set()
    actualizados = 0
    sin_match    = 0
    errores      = 0

    with engine.begin() as conn:
        for row in data_rows:
            if not any(row):
                continue
            try:
                id_pedido    = safe_int(row[1])
                agencia_carga= safe_str(row[4])
                vendedor     = safe_str(row[5])

                if not id_pedido:
                    continue
                if agencia_carga:
                    agencias.add(agencia_carga)

                result = conn.execute(text("""
                    UPDATE dartis_ventas
                    SET    vendedor     = :vendedor,
                           agencia_carga = :agencia_carga
                    WHERE  id_pedido = :id_pedido
                      AND  (vendedor IS NULL OR agencia_carga IS NULL)
                """), {
                    "vendedor": vendedor,
                    "agencia_carga": agencia_carga,
                    "id_pedido": id_pedido,
                })

                if result.rowcount > 0:
                    actualizados += result.rowcount
                else:
                    sin_match += 1

            except Exception as e:
                errores += 1
                print(f"  AVISO fila omitida: {e}")

        print(f"  Registros actualizados : {actualizados}")
        print(f"  Sin coincidencia       : {sin_match}")
        if errores:
            print(f"  Errores               : {errores}")

        nuevas_ag = sync_agencias(conn, agencias)
        if nuevas_ag:
            print("\n  Agencias nuevas agregadas:")
            for ag in nuevas_ag:
                print(f"    -> {ag}")
        else:
            print("  Agencias de carga: todas ya existian")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Importar/enriquecer dartis_ventas desde Excel de Dartis")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_import = sub.add_parser("import", help="Importar Ventas Recetas")
    p_import.add_argument("--file", required=True, help="Ruta al Excel de Ventas Recetas")

    p_enrich = sub.add_parser("enrich", help="Rellenar vendedor y agencia_carga desde Ventas clasico")
    p_enrich.add_argument("--file", required=True, help="Ruta al Excel de Ventas clasico")

    args = parser.parse_args()
    archivo = Path(args.file)
    if not archivo.exists():
        print(f"ERROR: Archivo no encontrado: {archivo}")
        sys.exit(1)

    print(f"\nArchivo: {archivo.name}\n")

    if args.cmd == "import":
        cmd_import(archivo)
    elif args.cmd == "enrich":
        cmd_enrich(archivo)

    print("\nListo.\n")


if __name__ == "__main__":
    main()
